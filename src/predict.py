import argparse
import json
import os
import sys
from pathlib import Path

import numpy as np
import onnxruntime as ort
from transformers import AutoTokenizer

from src.preprocess import build_input_text, price_band
from src.taxonomy import load_taxonomy

BACKBONE_DIR = Path("models/tokenizer")
ONNX_DIR = Path("models/onnx")
# Must match the CSV the models were trained against (the full shop grid,
# not the old 3833-leaf canonical one). Overridable via env for flexibility.
TAXONOMY_CSV = Path(os.getenv("TAXONOMY_CSV", "taxonomy_live.csv"))
MAX_LENGTH = 128
NEG_INF = -1e10


def _assemble_text(
    title: str,
    brand: str | None = None,
    description: str | None = None,
    price: float | None = None,
) -> str:
    """Mirror src.build_real_training_data.build_text_for_training:
    title | brand | price_band | description[:200], all passed through
    clean_text. Training never saw anything else, so inference must match.
    """
    extras = []
    pb = price_band(price)
    if pb:
        extras.append(pb)
    return build_input_text(
        title=title,
        brand=brand,
        description=description,
        extra_fields=extras,
    )


def _softmax(logits: np.ndarray) -> np.ndarray:
    e = np.exp(logits - logits.max(axis=-1, keepdims=True))
    return e / e.sum(axis=-1, keepdims=True)


def _mask_logits(logits: np.ndarray, valid_indices: list) -> np.ndarray:
    result = np.full_like(logits, NEG_INF)
    result[:, valid_indices] = logits[:, valid_indices]
    return result


def _load_sessions() -> dict:
    sessions = {}
    for stage in ("super", "parent", "leaf"):
        path = ONNX_DIR / f"{stage}_classifier.onnx"
        if not path.exists():
            print(json.dumps({"error": "missing_model", "stage": stage}), flush=True)
            sys.exit(1)
        sessions[stage] = ort.InferenceSession(str(path))
    return sessions


def predict_one(
    title: str,
    tokenizer,
    sessions: dict,
    taxonomy: dict,
    brand: str | None = None,
    description: str | None = None,
    price: float | None = None,
) -> dict:
    text = _assemble_text(title, brand=brand, description=description, price=price)
    if not text:
        raise ValueError("empty input after preprocessing")
    enc = tokenizer(
        [text], max_length=MAX_LENGTH, padding="max_length",
        truncation=True, return_tensors="np",
    )
    feeds = {
        "input_ids": enc["input_ids"].astype(np.int64),
        "attention_mask": enc["attention_mask"].astype(np.int64),
    }

    # Super stage
    super_logits = sessions["super"].run(["logits"], feeds)[0]
    super_probs = _softmax(super_logits)[0]
    super_idx = int(np.argmax(super_probs))
    try:
        super_name = taxonomy["idx_to_super"][str(super_idx)]
    except KeyError:
        print(json.dumps({"error": "taxonomy_key_error", "stage": "super", "index": super_idx}), flush=True)
        raise ValueError(f"unknown super index {super_idx}")
    super_conf = float(super_probs[super_idx])

    # Parent stage — masked to valid children of predicted super
    valid_parents = [
        taxonomy["parent_to_idx"][p]
        for p in taxonomy["super_to_parents"].get(super_name, [])
        if p in taxonomy["parent_to_idx"]
    ]
    parent_logits = sessions["parent"].run(["logits"], feeds)[0]
    if valid_parents:
        parent_logits = _mask_logits(parent_logits, valid_parents)
    parent_probs = _softmax(parent_logits)[0]
    parent_idx = int(np.argmax(parent_probs))
    try:
        parent_name = taxonomy["idx_to_parent"][str(parent_idx)]
    except KeyError:
        print(json.dumps({"error": "taxonomy_key_error", "stage": "parent", "index": parent_idx}), flush=True)
        raise ValueError(f"unknown parent index {parent_idx}")
    parent_conf = float(parent_probs[parent_idx])

    # Leaf stage — masked to valid children of predicted parent
    valid_leaves = [
        taxonomy["leaf_to_idx"][l]
        for l in taxonomy["parent_to_leaves"].get(parent_name, [])
        if l in taxonomy["leaf_to_idx"]
    ]
    leaf_logits = sessions["leaf"].run(["logits"], feeds)[0]
    if valid_leaves:
        leaf_logits = _mask_logits(leaf_logits, valid_leaves)
    leaf_probs = _softmax(leaf_logits)[0]
    leaf_idx = int(np.argmax(leaf_probs))
    try:
        leaf_name = taxonomy["idx_to_leaf"][str(leaf_idx)]
    except KeyError:
        print(json.dumps({"error": "taxonomy_key_error", "stage": "leaf", "index": leaf_idx}), flush=True)
        raise ValueError(f"unknown leaf index {leaf_idx}")
    leaf_conf = float(leaf_probs[leaf_idx])

    return {
        "text": title,
        "input_text": text,
        "super": super_name,
        "parent": parent_name,
        "leaf": leaf_name,
        "super_conf": round(super_conf, 4),
        "parent_conf": round(parent_conf, 4),
        "leaf_conf": round(leaf_conf, 4),
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--text", default=None, help="Single product name to classify")
    parser.add_argument("--brand", default=None, help="Optional brand (matches training format)")
    parser.add_argument("--description", default=None, help="Optional description (matches training format)")
    parser.add_argument("--price", default=None, help="Optional price (feeds price_band token)")
    parser.add_argument("--xlsx", default=None, help="Path to XLSX file for batch classification")
    parser.add_argument("--export", default=None, help="Path to write results XLSX (batch mode only)")
    args = parser.parse_args()

    try:
        price_val = float(args.price) if args.price not in (None, "") else None
    except ValueError:
        price_val = None

    if not BACKBONE_DIR.exists():
        print(json.dumps({"error": "missing_tokenizer"}), flush=True)
        sys.exit(1)

    try:
        tokenizer = AutoTokenizer.from_pretrained(str(BACKBONE_DIR))
        taxonomy = load_taxonomy(TAXONOMY_CSV)
        sessions = _load_sessions()
    except Exception as e:
        print(json.dumps({"error": "init_failed", "detail": str(e)}), flush=True)
        sys.exit(1)

    if args.text is not None:
        text = args.text.strip()
        if not text:
            print(json.dumps({"error": "empty_text"}), flush=True)
            sys.exit(1)
        try:
            result = predict_one(
                text, tokenizer, sessions, taxonomy,
                brand=args.brand, description=args.description, price=price_val,
            )
        except Exception as e:
            print(json.dumps({"error": "predict_failed", "text": text, "detail": str(e)}), flush=True)
            sys.exit(1)
        print(json.dumps(result), flush=True)

    elif args.xlsx is not None:
        _run_batch(args.xlsx, args.export, tokenizer, sessions, taxonomy)

    else:
        print(json.dumps({"error": "no_mode_selected"}), flush=True)
        sys.exit(1)


def _run_batch(xlsx_path: str, export_path, tokenizer, sessions: dict, taxonomy: dict):
    import openpyxl
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        print(json.dumps({"error": "xlsx_load_failed", "detail": str(e)}), flush=True)
        print(json.dumps({"status": "done", "rows": 0}), flush=True)
        return

    # Skip header row if first cell is a non-numeric string. When we detect a
    # header, map known column names so callers can include optional Brand /
    # Description / Price columns (in any order) and have them fed to the same
    # preprocessing the model was trained on.
    col_idx = {"title": 0, "brand": None, "description": None, "price": None}
    skipped_header = False
    if rows and rows[0][0] is not None:
        first = str(rows[0][0]).strip()
        if first and not first.replace(".", "").replace("-", "").isdigit():
            header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
            aliases = {
                "title": {"title", "product", "name", "product name", "продукт", "название"},
                "brand": {"brand", "бренд", "марка"},
                "description": {"description", "desc", "описание"},
                "price": {"price", "цена"},
            }
            for i, h in enumerate(header):
                for key, names in aliases.items():
                    if h in names and col_idx[key] is None:
                        col_idx[key] = i
            rows = rows[1:]
            skipped_header = True

    def _cell(row, idx):
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        return None if v is None else str(v).strip() or None

    def _price(row):
        if col_idx["price"] is None or col_idx["price"] >= len(row):
            return None
        v = row[col_idx["price"]]
        if v is None or v == "":
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    results = []
    sheet_row = 2 if skipped_header else 1  # 1-based Excel rows
    for row in rows:
        text = _cell(row, col_idx["title"]) or ""
        if not text:
            sheet_row += 1
            continue
        try:
            result = predict_one(
                text, tokenizer, sessions, taxonomy,
                brand=_cell(row, col_idx["brand"]),
                description=_cell(row, col_idx["description"]),
                price=_price(row),
            )
            result["row"] = sheet_row
            print(json.dumps(result), flush=True)
            results.append(result)
        except Exception as e:
            print(json.dumps({"error": "predict_failed", "row": sheet_row, "text": text, "detail": str(e)}), flush=True)
        sheet_row += 1

    if export_path and results:
        try:
            wb_out = openpyxl.Workbook()
            ws_out = wb_out.active
            ws_out.append(["Product", "Super", "Parent", "Leaf", "Confidence"])
            for r in results:
                ws_out.append([
                    r["text"], r["super"], r["parent"], r["leaf"],
                    f"{r['leaf_conf'] * 100:.1f}%",
                ])
            wb_out.save(export_path)
            print(json.dumps({"status": "export_done", "path": export_path, "rows": len(results)}), flush=True)
        except Exception as e:
            print(json.dumps({"error": "export_failed", "detail": str(e)}), flush=True)

    print(json.dumps({"status": "done", "rows": len(results)}), flush=True)


if __name__ == "__main__":
    main()
